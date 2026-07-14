import sys
import openpyxl
import os

def analyze():
    file_path = '/Users/ryanrichard/projecont/Rayo/temp/braga/planilha-importacao-dealer.xlsm'
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return

    print(f"Reading file: {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=False, read_only=False)
    
    for sheet_name in wb.sheetnames:
        print(f"\n--- Sheet: {sheet_name} ---")
        sheet = wb[sheet_name]
        
        for i, row in enumerate(sheet.iter_rows(values_only=False)):
            if i > 5:  # Only look at the first few rows for layout
                break
            
            row_data = []
            for cell in row:
                val = cell.value
                if val is not None:
                    # if it's a formula, print it. If not, print value
                    if isinstance(val, str) and val.startswith('='):
                        row_data.append(f"{cell.column_letter}: {val}")
                    else:
                        row_data.append(f"{cell.column_letter}: {val}")
            if row_data:
                print(f"Row {i+1}: {' | '.join(row_data)}")
                
if __name__ == '__main__':
    analyze()
